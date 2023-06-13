import tkinter as tk
from tkinter import * 
from  tkinter import filedialog, Text 
import customtkinter
import os 
import pandas as pd 
import csv 
import openpyxl
from tkinter import ttk
from functools import partial
from openpyxl.workbook import Workbook 
from openpyxl import load_workbook
import hashlib
import re
from subprocess import call
import time
import sys


customtkinter.set_appearance_mode("Dark")  # Modes: "System" (standard), "Dark", "Light"
customtkinter.set_default_color_theme("blue")  # Themes: "blue" (standard), "green", "dark-blue"

root = customtkinter.CTk()

# this will create a label widget
l1 = customtkinter.CTkLabel(root, text = "")
l2 = customtkinter.CTkLabel(root, text = "")

# grid method to arrange labels in respective
# rows and columns as specified
l1.grid(row = 0, column = 0, sticky = W, pady = 2)
l2.grid(row = 3, column = 0, sticky = W, pady = 2)

def change_appearance_mode_event(root, new_appearance_mode: str):
    customtkinter.set_appearance_mode(new_appearance_mode)


root.title("paTTeRn")
root.geometry(f"{1100}x{580}")

root.tabview = customtkinter.CTkTabview(root, width=500,height=200)
root.tabview.grid(row=0, column=1, padx=(10, 0), pady=(5, 0), sticky="nsew")
#root.tabview.add("CTkTabview")
root.tabview.add("Manually Input Data")
root.tabview.add("Pull Data with Aqua (Beta)")
#root.tabview.tab("CTkTabview").grid_columnconfigure(0, weight=1)  # configure grid of individual tabs
root.tabview.tab("Manually Input Data").grid_columnconfigure(0, weight=1)
root.tabview.tab("Pull Data with Aqua (Beta)").grid_columnconfigure(0, weight=1)

root.appearance_mode_label = customtkinter.CTkLabel(root, text="Appearance Mode:", anchor="w")
root.appearance_mode_label.grid(row=5, column=4, padx=20, pady=(10, 0))
root.appearance_mode_optionemenu = customtkinter.CTkOptionMenu(root, values=["Light", "Dark", "System"],
                                                               command=change_appearance_mode_event)
root.appearance_mode_optionemenu.grid(row=6, column=4, padx=20, pady=(10, 10))

def run_aqua():
    print("Initiating Aqua Pull...")
    TP_wildcard = input("What test program are you pulling for (i.e. give TP wildcard): ")
    ndays_wildcard = input("How many days data do you want: ")
    global current_directory
    current_directory = os.getcwd()
    print(current_directory)

   
    aqua = ['\\\\GER.corp.intel.com\ec\proj\ha\stav\DIS_Downloads\AquaHbase\AquaCMDClient\Client\AquaCmdLine.exe','-aquaServer', 'GER',
     '-reportpath', "cathalba\Python\ScanFi_00", '-outputFileName', str(current_directory)+"\output.csv", '-sendmail', '-lastndaysteststart', str(ndays_wildcard), '-programNames', str(TP_wildcard)
]

    exitprogram = "0"

    if os.path.isfile('output.csv') and exitprogram != "x" :
        del_outputfile = input("output.csv file already exists in directory. Do you want to delete and proceed? (y/n) ")
        print(os.path)
        if del_outputfile == "y":
            try:
                os.remove("output.csv")
            except IOError:
                print("***ERROR*** \nIs the output.csv file currently open? \n Exiting...")
                time.sleep(1)
                os._exit(0)
   
        else:
            print("Exiting Script")
            exit()

    else:
        print("No output.csv detected in current directory. Proceeding with JMP Query")

    print(aqua)
    call(aqua)
    if not os.path.isfile('output.csv'):
      print("\n")
      exitprogram = input("->>> Your search came empty! <<<- Enter \"x\" to exit, any other key to continue: ")

def Excel_File_Select():

    #Choose Excel data
    global FailingPatterns
    global data_directory

    #Select direcotry where data is located + where output plist will be saved
    data_directory = tk.filedialog.askdirectory(title = " Choose a directory")
    os.chdir(data_directory)
    print("Current working directory: {0}".format(os.getcwd()))
    FailingPatterns = tk.filedialog.askopenfilename(parent=root,title = "Choose failing pats")
    print(root.tk.splitlist(FailingPatterns))

def read_Plist():
    global plist
    plist_text = textbox.get("1.0",END)
    print("Plist = " + plist_text)
    print("Successfully read Plist")
    temp_file = data_directory + "\\plist_temp.txt"
    with open(temp_file, "w") as file:
        file.write(plist_text)

def TCScore_Output_Trim():
    
    global tc_score_tuples
    tc_score_output = tk.filedialog.askopenfilename(parent=root, title = "Select TCSCore output file", filetypes=(("CSV files", "*.csv"), ("Any file", "*")))
    df = pd.read_csv(tc_score_output, on_bad_lines='skip')

    patterns = df["Name"].str[1:8]

    df_patterns = pd.DataFrame({"Patterns": patterns})

    file_path = os.path.dirname(tc_score_output)
    df_patterns.to_csv(file_path + "\\patterns.csv", index=False)
    tc_score_tuples = file_path + "\\patterns.csv"
    print("I've successfully saved your TCScore output as a list of tuples in a csv file.")

    workbook = openpyxl.Workbook()
    worksheet = workbook.create_sheet("Sheet1")

    with open(file_path + "\\patterns.csv", "r") as csv_file: 
        data = csv.reader(csv_file)
        # Add the data from the CSV file to the Excel worksheet
        for row in data:
            worksheet.append(row)
    
    # Save the Excel workbook
    workbook.save(file_path + "\\patterns.xlsx")
    print("I've successfully updated your TCScore output as an excel file list of tuples which can now be read by the paTTeRn script.")
    print("File has been saved to " + str(file_path))



#Main Script Body
def run_Script():
    try:
        wb = load_workbook(FailingPatterns)
    except IOError:
        print("***ERROR*** \nIs the pattern excel file currently open? \nExiting...")
        time.sleep(1)
        os._exit(0)

    ws = wb.active
    
    xl = pd.ExcelFile(FailingPatterns)
    df = xl.parse("Patterns") #or Sheet1? Need to fix 
    
    x = df['Patterns']
    temp_file = data_directory + "\\plist_temp.txt"
    plist = open(temp_file, "r")
    
    rows = len(df.axes[0])
    
    print("Number of Rows: ", rows)
    
    outfile_name = data_directory + "\\TTR_Plist.plist"
    temp_file = data_directory + "\\temp.txt"
    
    temp = open(temp_file, "w")
    #plist = open(Plist_path, "r")
    
    #regex commands
    #Asks for user input about if the template is the standard FAST/IPSearch. If not then we are using VMIN which prints pattern data in a different way (i.e. just the tuple)
    

    #template = input("Does this vmin use FAST/IPSearch templates?: y/n ")
    
    #if template == 'on':
    if selected_option.get() == "Vmin Template":
        regex_tuple = "Pat (d|g)(0|1|2)(\d+)"
        regex_group = 3
        print("Regex search = " + str(regex_group))

    elif selected_option.get() == "IPSearch Template":
        regex_tuple = "Pat (d|g)(0|1|2)(\d+)"
        regex_group = 3
        print(regex_group)
        print(regex_tuple)
    else:
        regex_tuple = "Pat (\w+);"
        regex_group = 1
        print("I am printing for the FAST template")
    
    
    #if template2.get() == "on" and template.get() == "on":
    #    print("TCScore output and VMin template? That's impossible!")

    if (selected_option.get() == "IPSearch Template") & (TC_Score.get() == "on"):
        print("I will reduce your plist using the redundant pattern list from TCScore")
        with open(temp_file, "w") as temp:
            for line in plist:
                try:
                    tuple = re.search(regex_tuple, line).group(regex_group)
                    print("my tuple =" + str(tuple))
                except AttributeError:
                    tuple = re.search(regex_tuple, line)
                if "GlobalPList" in line or "rpl_pst" in line or "_stfinit_" in line or "PList" in line or "}" in line: 
                    temp.write(line)
                elif not line.strip():
                    print("This line is empty, I'll do nothing.")
                    temp.write(line)
                elif x.astype(str).str.contains(str(tuple)).any():
                    print("This is a redundant tuple, remove it")
                    line = "#" + line
                    print(str(line) + "I will write this line to the temp file")
                    temp.write(line)

                else: 
                    temp.write(line)
    elif (selected_option.get() == "IPSearch Template") & (TC_Score.get() == "off"):
        print("I will reduce your plist using the failing pattern list you have inputted manually")
        with open(temp_file, "w") as temp:
            for line in plist:
                try:
                    tuple = re.search(regex_tuple, line).group(regex_group)
                    print("my tuple =" + str(tuple))
                except AttributeError:
                    tuple = re.search(regex_tuple, line)
                if "GlobalPList" in line or "rpl_pst" in line or "_stfinit_" in line or "PList" in line or "}" in line: 
                    temp.write(line)
                elif not line.strip():
                    print("This line is empty, I'll do nothing.")
                    temp.write(line)
                elif x.astype(str).str.contains(str(tuple)).any():
                    print("This is a failing tuple, we need to keep it")
                    temp.write(line)
                else: 
                    line = "#" + line
                    temp.write(line)
            temp.close() 

    else:
        print("I'm running else")
        print(TC_Score.get())
        for line in plist:
            try:
                #print("line = " + str(line))
                tuple = re.search(regex_tuple, line).group(regex_group)
                print("my tuple =" + str(tuple))
            except AttributeError:
                tuple = re.search(regex_tuple, line)
            if "GlobalPList" in line or "rpl_pst" in line or "_stfinit_" in line or "PList" in line or "}" in line: 
                temp.write(line)
            elif not line.strip():
                print("This line is empty, I'll do nothing.")
                temp.write(line)
            elif x.astype(str).str.contains(str(tuple)).any():
                print("This is a failing tuple, keep it")
                temp.write(line)
            else: 
                line = "#" + line
                temp.write(line)
                print("This tuple never fails, remove it")
        temp.close()          
    
    
    lines_seen = set()
    outfile = open(outfile_name, "w")
    for line in open(temp_file, "r"): 
        #if line not in lines_seen:
        #    outfile.write(line)
        #    lines_seen.add(line)
        outfile.write(line)
    outfile.close()

    #if os.path.exists("temp.txt"):
    #  os.remove("temp.txt")
    #  print("temp file deleted")
    #else:
    #  print("temp file does not exist")



frame = customtkinter.CTkFrame(master=root,
                               width=15,
                               height=18,
                               corner_radius=10)
frame.grid(row = 0, column = 0, columnspan = 3, rowspan = 3, pady = 2)

textbox = customtkinter.CTkTextbox(root,width=600,height=200,corner_radius=10)
#textbox.grid(row=0, column=0)
textbox.grid(row = 1, column = 1, columnspan = 3, rowspan = 2, pady = 5)
#scrollbar4 = customtkinter.CTkScrollbar(root, command=textbox.yview)
#scrollbar4.grid(row=0, column=5, sticky="nsew")
#textbox.configure(yscrollcommand=scrollbar4.set)

button = customtkinter.CTkButton(root.tabview.tab("Manually Input Data"), text="Open Excel File", command=Excel_File_Select)
button.grid(row = 0, column = 0, columnspan = 2, pady = 5)

button2 = customtkinter.CTkButton(master=root, text="Read Plist", command=read_Plist)
button2.grid(row = 4, column = 2, columnspan = 2, pady = 5)

button3 = customtkinter.CTkButton(master=root, text="Run Script", command=run_Script)
button3.grid(row = 5, column = 2, columnspan = 2, pady = 5)

button4 = customtkinter.CTkButton(root.tabview.tab("Manually Input Data"), text="Process TCScore Output", command=TCScore_Output_Trim)
button4.grid(row = 1, column = 0, columnspan = 2, pady = 5)

Aqua_button = customtkinter.CTkButton(root.tabview.tab("Pull Data with Aqua (Beta)"), text="Start Aqua Pull", command=run_aqua)
Aqua_button.grid(row = 0, column = 0, columnspan = 2, pady = 5)

#dialog = customtkinter.CTkInputDialog(root.tabview.tab("Pull Data with Aqua (Beta)"), text="Type in a number:", title="Test")
#dialog.grid(row = 1, column = 0, columnspan = 2, pady = 5)

template = tk.StringVar(value="on")
template2 = tk.StringVar(value="on")
template3 = tk.StringVar(value="on")
selected = tk.StringVar(value="on")

#def checkbox_event():
#    print("checkbox toggled, current value:", template.get())
#
#def checkbox2_event():
#    print("TCScore Checkbox toggled, current value:", template.get())
#checkboxes = [
##checkbox = customtkinter.CTkCheckBox(master=root, text="VMin Template?", command=checkbox_event,
#                                     variable=template, onvalue="on", offvalue="off")
#checkbox.grid(row = 5, column = 0, pady = 2)
#
#checkbox2 = customtkinter.CTkCheckBox(master=root, text="IPSearch? (requires TC score output)", command=checkbox2_event,
#                                     variable=template2, onvalue="on", offvalue="off")
#checkbox2.grid(row = 4, column = 0, pady = 2)
#
#checkbox3 = customtkinter.CTkCheckBox(master=root, text="FAST Template?",
#                                     variable=template3, onvalue="on", offvalue="off")
#checkbox3.grid(row = 6, column = 0, pady = 2)

#customtkinter.CTkCheckBox(master=root, text="VMin Template?", command=checkbox_event,
#                                     variable=selected, onvalue="VMin Template"),
#customtkinter.CTkCheckBox(master=root, text="IPSearch? (requires TC score output)", command=checkbox2_event,
#                                     variable=selected, onvalue="IPSearch Template"),
#customtkinter.CTkCheckBox(master=root, text="FAST Template?",
#                                     variable=selected, onvalue="FAST Template")
#]
#
#for checkbox in checkboxes:
#    checkbox.grid(column = 0, pady = 2)

# Create a variable to store the selected option
selected_option = customtkinter.StringVar()
TC_Score = customtkinter.StringVar(value="off")

# Create the radio buttons
option1 = customtkinter.CTkRadioButton(master=root, text="VMin Template", variable=selected_option, value="Vmin Template")
option1.grid(row = 4, column = 0, pady=10, padx=20, sticky="n")
option2 = customtkinter.CTkRadioButton(master=root, text="IPSearch", variable=selected_option, value="IPSearch Template")
option2.grid(row = 5, column = 0, pady=10, padx=20, sticky="n")
option4 = customtkinter.CTkCheckBox(master=root, text="Are you using TC Score?", variable=TC_Score, onvalue="on", offvalue="off")
option4.grid(row = 5, column = 1, pady=10, padx=20, sticky="n")
option3 = customtkinter.CTkRadioButton(master=root, text="FAST Template", variable=selected_option, value="FAST Template")
option3.grid(row = 6, column = 0, pady=10, padx=20, sticky="n")



root.mainloop()

